"""
Module to generate an Excel documentation file from database schema information.

This module uses the openpyxl library to create a structured and well-formatted
Excel workbook, including a summary sheet and individual sheets for each table.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from db_utils import run_mysql

def generate_excel_doc(output_file, db_container, db_user, db_pass, db_name, tables, fk_map):
    """
    Generates a comprehensive Excel documentation file for the given database schema.

    Args:
        output_file (str): The path and name for the output Excel file.
        db_container (str): The name of the Docker container.
        db_user (str): The database username.
        db_pass (str): The database password.
        db_name (str): The name of the database.
        tables (list): A list of table names to document.
        fk_map (dict): A dictionary mapping foreign key columns to their referenced tables/columns.
    """
    wb = Workbook()
    summary_ws = wb.active
    summary_ws.title = "Tables"

    # Define common styles
    header_fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
    table_title_fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
    border_style = Border(
        left=Side(border_style="thin", color="000000"),
        right=Side(border_style="thin", color="000000"),
        top=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="thin", color="000000")
    )

    # Add title and date to the summary sheet
    summary_ws["A1"] = "📖 PMS Database Documentation"
    summary_ws["A1"].font = Font(bold=True, size=14)
    summary_ws.merge_cells("A1:B1")

    summary_ws["A2"] = f"Documentation Date: {datetime.now().strftime('%B %d, %Y %I:%M %p')}"
    summary_ws.merge_cells("A2:B2")
    summary_ws["A2"].font = Font(italic=True, size=11, color="555555")

    # Add header for the table list
    summary_ws.append([])
    summary_ws.append(["Table Name", "Go to Sheet"])
    for col in ("A", "B"):
        cell = summary_ws[col + "4"]
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border_style

    # Keep track of table -> sheet name mapping
    table_to_sheet = {}

    # Loop through all tables and create a dedicated sheet for each
    for idx, table in enumerate(tables, start=1):
        print(f"  → Processing {table}")
        
        # Ensure sheet name is no more than 31 characters
        sheet_name = table[:31]
        ws = wb.create_sheet(title=sheet_name)
        table_to_sheet[table] = sheet_name

        # Add table title
        ws.merge_cells("A1:G1")
        title_cell = ws["A1"]
        title_cell.value = f"Physical Table Name: {table}"
        title_cell.font = Font(bold=True, size=14, color="FFFFFF")
        title_cell.fill = table_title_fill
        title_cell.alignment = Alignment(horizontal="left", vertical="center")

        # Add header row for table details
        headers = ["Physical Column Name", "Type", "Primary Key", "Allow NULL", "Default Value", "Extra", "Comment"]
        ws.append([""] * len(headers))  # spacer row
        ws.append(headers)

        # Style headers
        header_row = 3
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=header_row, column=col)
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border_style

        # Run DESCRIBE on the current table
        try:
            result = run_mysql(db_container, db_user, db_pass, db_name, f"DESCRIBE `{table}`;")
        except Exception as e:
            print(f"Warning: Could not describe table {table}. Error: {e}")
            continue

        row_idx = header_row + 1
        for line in result.split("\n"):
            if not line:
                continue
            parts = line.split("\t")
            
            # Handle cases where some fields might be empty
            while len(parts) < 6:
                parts.append("")

            col, col_type, nullable, key, default, extra = parts[:6]

            # Interpret key types
            if key == "PRI":
                key_val = "PRI"
            elif key == "MUL":
                key_val = "FK"
            elif key == "UNI":
                key_val = "UNI"
            else:
                key_val = "-"

            # Interpret default values
            if nullable == "NO" and default == "NULL":
                default_val = ""
            elif default == "" or default is None:
                default_val = "-"
            else:
                default_val = default

            row_data = [col, col_type, key_val, nullable, default_val, extra if extra else "-", ""]
            ws.append(row_data)

            # Apply borders and check for foreign keys
            for c in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=c).border = border_style

            if (table, col) in fk_map:
                ref_table, ref_col = fk_map[(table, col)]
                target_sheet = table_to_sheet.get(ref_table, ref_table[:31])
                link_cell = ws.cell(row=row_idx, column=7)  # Comment column
                link_cell.value = f"FK → {ref_table}.{ref_col}"
                link_cell.hyperlink = f"#{target_sheet}!A1"
                link_cell.style = "Hyperlink"

            row_idx += 1

        # Freeze the header row for easy scrolling
        ws.freeze_panes = ws["A4"]

        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max_length + 2

        # Add a "Home" button to return to the summary sheet
        home_cell = ws["I1"]
        home_cell.value = "🏠 Home"
        home_cell.hyperlink = "#Tables!A1"
        home_cell.style = "Hyperlink"
        home_cell.font = Font(bold=True)

        # Add this table to the summary sheet with a hyperlink
        summary_ws.append([table, "Go"])
        summary_ws.cell(row=idx + 4, column=2).hyperlink = f"#{sheet_name}!A1"
        summary_ws.cell(row=idx + 4, column=2).style = "Hyperlink"

    # Freeze the summary header
    summary_ws.freeze_panes = "A5"

    # Auto-adjust summary column widths
    for col in summary_ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        summary_ws.column_dimensions[col_letter].width = max_length + 2

    # Save the final workbook
    wb.save(output_file)
